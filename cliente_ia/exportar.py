"""
MV Cliente IA · exportación
============================
La lista sirve cuando entra al CRM o al secuenciador de correos, así que la
corrida sale en CSV (universal) y en XLSX cuando hay openpyxl.

Formato de números: el estándar del proyecto —score con un decimal, sin
notación científica— y una fila por decisor, que es la unidad con la que
trabaja quien manda los correos.
"""
from __future__ import annotations

import csv
import io
import re
from pathlib import Path

from . import rutas
from .modelos import Corrida

COLUMNAS = [
    "prioridad", "nivel", "pais", "score", "empresa", "dominio", "sector",
    "ciudad", "empleados", "decisor", "cargo", "seniority", "email", "idioma",
    "asunto", "cuerpo", "seguimiento", "linkedin", "linkedin_nota",
    "video_url", "landing_url", "senales", "campana", "sintetico",
]


# Excel y Google Sheets tratan la celda como FÓRMULA si arranca con alguno de
# estos. El texto de las celdas no lo escribe quien descarga: sale del <title>
# y los <h1> de sitios ajenos (fase 1 y el rastreo de contactos) y del JSON
# del modelo. Un sitio que el cliente investigue puede llamarse
# `=cmd|'/c calc'!A0` y Excel lo ejecuta al abrir el archivo — y este export
# existe justamente para abrirse en el CRM del cliente.
_ARRANQUES_PELIGROSOS = ("=", "+", "-", "@", "\t", "\r", "\n")


def _sanear_celda(valor):
    """Antepone un apóstrofo a lo que Excel interpretaría como fórmula. El
    apóstrofo no se ve en la celda: le dice a Excel «esto es texto».

    Sólo toca strings: los números (score, empleados, prioridad) viajan como
    int/float y un negativo legítimo no se rompe."""
    if isinstance(valor, str) and valor[:1] in _ARRANQUES_PELIGROSOS:
        return "'" + valor
    return valor


def filas(corrida: Corrida) -> list[dict]:
    """Una fila por decisor, con su correo si la fase 6 lo escribió.

    Los valores salen ya saneados contra inyección de fórmulas: es el único
    lugar por el que pasan tanto el CSV como el XLSX."""
    prospectos = {p.id: p for p in corrida.prospectos}
    emails = {e.decisor_id: e for e in corrida.emails}
    campanas = {c.id: c for c in corrida.campanas}
    salida: list[dict] = []
    for d in corrida.decisores:
        p = prospectos.get(d.prospecto_id)
        if not p:
            continue
        e = emails.get(d.id)
        c = campanas.get(p.campana_id)
        salida.append({
            "prioridad": p.prioridad,
            "nivel": p.nivel,
            "pais": p.pais,
            "score": round(d.score, 1),
            "empresa": p.nombre,
            "dominio": p.dominio,
            "sector": p.sector,
            "ciudad": p.ciudad,
            "empleados": p.empleados,
            "decisor": d.nombre,
            "cargo": d.cargo,
            "seniority": d.seniority,
            "email": d.email,
            "idioma": d.idioma,
            "asunto": e.asunto if e else "",
            "cuerpo": e.cuerpo if e else "",
            "seguimiento": e.seguimiento if e else "",
            # LinkedIn y los enlaces viajan en el CSV para poder pegarlos
            # directo en el secuenciador o en la herramienta de LinkedIn.
            "linkedin": e.linkedin if e else "",
            "linkedin_nota": e.linkedin_nota if e else "",
            "video_url": e.video_url if e else "",
            "landing_url": e.landing_url if e else "",
            "senales": " | ".join(p.senales),
            "campana": c.nombre if c else p.campana_id,
            "sintetico": "sí" if (d.sintetico or p.sintetico) else "no",
        })
    return [{k: _sanear_celda(v) for k, v in fila.items()} for fila in salida]


def _destino(corrida: Corrida, ext: str) -> Path:
    """La ruta del archivo a escribir, con el id validado igual que en
    `almacen.ruta_de` y el dominio saneado.

    Hace falta porque `POST /api/exportar/{csv,xlsx}` reconstruye la corrida
    desde el cuerpo del pedido: con `id="../../../pwned"` el nombre de
    archivo se escapaba del directorio de exports y el XLSX terminaba en
    cualquier ruta escribible por el proceso."""
    if not corrida.id or not re.fullmatch(r"[A-Za-z0-9_-]{1,64}", corrida.id):
        raise ValueError(f"id de corrida inválido: {corrida.id!r}")
    # El dominio SÍ se limpia en vez de rechazarse: es texto que escribe el
    # usuario y un punto o una barra de más no son un ataque, sólo un nombre
    # de archivo feo. El id, que es generado, es el que se valida duro.
    dominio = re.sub(r"[^A-Za-z0-9._-]", "_", corrida.dominio or "")[:80].strip("._") or "corrida"
    return rutas.dir_exports() / f"{dominio}_{corrida.id}{ext}"


def a_csv(corrida: Corrida) -> str:
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=COLUMNAS, lineterminator="\n")
    w.writeheader()
    w.writerows(filas(corrida))
    return buf.getvalue()


def guardar_csv(corrida: Corrida, destino: Path | None = None) -> Path:
    destino = destino or _destino(corrida, ".csv")
    destino.parent.mkdir(parents=True, exist_ok=True)
    # utf-8-sig: sin el BOM, Excel en Windows abre los acentos rotos.
    destino.write_text(a_csv(corrida), encoding="utf-8-sig")
    return destino


def guardar_xlsx(corrida: Corrida, destino: Path | None = None) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:                                # pragma: no cover
        raise RuntimeError("Falta openpyxl (pip install openpyxl)") from e

    destino = destino or _destino(corrida, ".xlsx")
    destino.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Prospectos"
    encabezado = Font(bold=True, color="FFFFFF")
    fondo = PatternFill("solid", start_color="0E1628")      # navy de la marca
    ws.append([c.upper() for c in COLUMNAS])
    for celda in ws[1]:
        celda.font = encabezado
        celda.fill = fondo
        celda.alignment = Alignment(horizontal="left")
    for fila in filas(corrida):
        ws.append([fila[c] for c in COLUMNAS])
    anchos = {"empresa": 34, "sector": 30, "decisor": 24, "cargo": 28, "email": 34,
              "asunto": 46, "cuerpo": 70, "seguimiento": 60, "linkedin": 70,
              "linkedin_nota": 46, "video_url": 40, "landing_url": 40,
              "senales": 60, "campana": 34}
    for i, col in enumerate(COLUMNAS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = anchos.get(col, 12)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    resumen = wb.create_sheet("Resumen")
    r = corrida.resumen()
    # Estos no pasan por `filas()`, así que se sanean acá uno por uno.
    resumen.append(["Dominio", _sanear_celda(corrida.dominio)])
    resumen.append(["Corrida", _sanear_celda(corrida.id)])
    resumen.append(["Modo", _sanear_celda(corrida.modo)])
    resumen.append(["Estado", _sanear_celda(corrida.estado)])
    resumen.append([])
    resumen.append(["Prospectos por ola", ""])
    for nivel, valor in r["prospectos_por_nivel"].items():
        resumen.append([nivel, valor])
    resumen.append([])
    resumen.append(["Correos por idioma", ""])
    for idioma, valor in r["emails_por_idioma"].items():
        resumen.append([idioma, valor])
    resumen.column_dimensions["A"].width = 24
    resumen.column_dimensions["B"].width = 40

    wb.save(destino)
    return destino
